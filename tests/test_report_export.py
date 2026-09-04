"""Contract: qa/contracts/report-export.md RE1-RE5."""

from __future__ import annotations

import base64
from pathlib import Path

import pytest
from openpyxl import load_workbook

from autotester.schema.case import Case
from autotester.schema.enums import (
    Action,
    CaseClass,
    CaseKind,
    EvidenceKind,
    Outcome,
    Result,
)
from autotester.schema.flowspec import Step
from autotester.schema.project import Project
from autotester.schema.run import Evidence, RawResult
from autotester.schema.verdict import Verdict
from autotester.stages import report_export
from autotester.store import ProjectStore

RUN_ID = "run-test123"
PNG_BYTES = base64.b64decode(  # a real, tiny, valid 1x1 PNG
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


def _seed(tmp_path: Path, *, outcome: Outcome, result: Result) -> ProjectStore:
    store = ProjectStore("demo", tmp_path)
    store.save_project(
        Project(slug="demo", name="Demo", base_url="https://demo.test",
                allowed_domains=["demo.test"])
    )
    case = Case(
        project="demo", flow_id="flow-login", kind=CaseKind.BEST, case_class=CaseClass.HAPPY,
        title="Login works",
        steps=[Step(order=1, action=Action.NAVIGATE, target="/login")],
    )
    store.add_case(case)
    run_dir = store.paths.run_dir(RUN_ID)
    run_dir.mkdir(parents=True)
    (run_dir / "01-step01-navigate.png").write_bytes(PNG_BYTES)
    raw = RawResult(
        case_id=case.id, outcome=outcome, duration_s=1.23,
        error=None if outcome is Outcome.COMPLETED else "TimeoutError: boom",
        evidence=[Evidence(kind=EvidenceKind.SCREENSHOT, path="01-step01-navigate.png",
                            step_order=1, label="step01-navigate")],
    )
    store.save_result(RUN_ID, raw)
    store.save_verdict(RUN_ID, Verdict(
        run_id=RUN_ID, case_id=case.id, result=result, criteria_met=1, criteria_total=1,
        scoreboard="Criteria 1/1 met.", grader_provider="gemini",
    ))
    return store


def test_export_excel_has_one_row_per_case(tmp_path: Path) -> None:
    _seed(tmp_path, outcome=Outcome.COMPLETED, result=Result.PASS)

    out = report_export.export_excel("demo", RUN_ID, tmp_path / "out.xlsx", tmp_path)

    wb = load_workbook(out)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "Case"
    assert rows[1][0] == "Login works"
    assert rows[1][4] == "PASS"  # Result column


def test_export_excel_defaults_to_the_latest_run(tmp_path: Path) -> None:
    _seed(tmp_path, outcome=Outcome.COMPLETED, result=Result.PASS)

    out = report_export.export_excel("demo", None, tmp_path / "out.xlsx", tmp_path)

    wb = load_workbook(out)
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[1][0] == "Login works"


def test_export_excel_raises_a_clear_error_with_no_runs(tmp_path: Path) -> None:
    ProjectStore("demo", tmp_path).save_project(
        Project(slug="demo", name="Demo", base_url="https://demo.test",
                allowed_domains=["demo.test"])
    )
    with pytest.raises(ValueError, match="no runs"):
        report_export.export_excel("demo", None, tmp_path / "out.xlsx", tmp_path)


def test_export_html_embeds_the_screenshot_and_is_self_contained(tmp_path: Path) -> None:
    _seed(tmp_path, outcome=Outcome.COMPLETED, result=Result.PASS)

    out = report_export.export_html("demo", RUN_ID, tmp_path / "out.html", tmp_path)

    html = out.read_text(encoding="utf-8")
    assert "Login works" in html
    assert "PASS" in html
    assert "data:image/png;base64," in html  # RE3: embedded, not a file reference
    assert "step01-navigate" in html


def test_export_html_shows_the_error_for_an_errored_case(tmp_path: Path) -> None:
    _seed(tmp_path, outcome=Outcome.ERRORED, result=Result.INCONCLUSIVE)

    out = report_export.export_html("demo", RUN_ID, tmp_path / "out.html", tmp_path)

    html = out.read_text(encoding="utf-8")
    assert "TimeoutError: boom" in html
    assert "INCONCLUSIVE" in html
